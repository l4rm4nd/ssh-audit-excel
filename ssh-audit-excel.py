import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse

# Setup argparse to get the directory path from command line
parser = argparse.ArgumentParser(description='Process JSON files for SSH auditing.')
parser.add_argument('-d', '--dir', required=True, help='Directory containing the JSON files')

args = parser.parse_args()
json_dir = args.dir

# Create lists to hold the detailed data
detailed_data = []

# Custom severity mapping based on specific findings
custom_severity_mapping = {
    # 'using weak cipher mode': 'low',
    # 'using broken SHA-1 hash algorithm': 'medium',
    # 'using small 1024-bit modulus': 'medium',
    # 'using elliptic curves that are suspected as being backdoored by the U.S. National Security Agency': 'high',
    # 'using broken & deprecated 3DES cipher': 'medium',
    # 'using weak & deprecated CAST cipher': 'low',
    # 'using weak & deprecated Blowfish cipher': 'low',
    # 'using small 64-bit block size': 'medium',
    # Add any additional mappings as needed
}

# Iterate over all JSON files in the directory
for filename in os.listdir(json_dir):
    if filename.endswith('.json'):
        with open(os.path.join(json_dir, filename), 'r') as file:
            try:
                data = json.load(file)
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON in file {filename}: {e}")
                continue
            
            # Extracting relevant data for the detailed tab
            target = data.get('target', '')
            
            # Handle CVEs - All CVE findings set to 'info'
            if 'cves' in data:
                for cve in data['cves']:
                    detailed_data.append({
                        'Target': target,
                        'Category': 'CVE',
                        'Item': cve['name'],
                        'Description': cve['description'],
                        'Audit': 'info',
                    })
            
            # Handle Encryption Algorithms
            for enc in data.get('enc', []):
                algorithm = enc['algorithm']
                for severity, notes in enc.get('notes', {}).items():
                    for note in notes:
                        custom_severity = custom_severity_mapping.get(note, severity)  # Use custom severity if defined
                        detailed_data.append({
                            'Target': target,
                            'Category': 'Encryption',
                            'Item': algorithm,
                            'Description': note,
                            'Audit': custom_severity,
                        })
            
            # Handle KEX Algorithms
            for kex in data.get('kex', []):
                algorithm = kex['algorithm']
                for severity, notes in kex.get('notes', {}).items():
                    for note in notes:
                        custom_severity = custom_severity_mapping.get(note, severity)  # Use custom severity if defined
                        detailed_data.append({
                            'Target': target,
                            'Category': 'KEX',
                            'Item': algorithm,
                            'Description': note,
                            'Audit': custom_severity,
                        })
            
            # Handle Key Algorithms
            for key in data.get('key', []):
                algorithm = key['algorithm']
                for severity, notes in key.get('notes', {}).items():
                    for note in notes:
                        custom_severity = custom_severity_mapping.get(note, severity)  # Use custom severity if defined
                        detailed_data.append({
                            'Target': target,
                            'Category': 'Key',
                            'Item': algorithm,
                            'Description': note,
                            'Audit': custom_severity,
                        })
            
            # Handle MAC Algorithms
            for mac in data.get('mac', []):
                algorithm = mac['algorithm']
                for severity, notes in mac.get('notes', {}).items():
                    for note in notes:
                        custom_severity = custom_severity_mapping.get(note, severity)  # Use custom severity if defined
                        detailed_data.append({
                            'Target': target,
                            'Category': 'MAC',
                            'Item': algorithm,
                            'Description': note,
                            'Audit': custom_severity,
                        })

# Create a DataFrame for the detailed data and remove duplicates
df_detailed = pd.DataFrame(detailed_data).drop_duplicates()

# Generate the filename with the current date and time
current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'SSH_Auditing_Results_{current_time}.xlsx'

# Create an Excel writer object
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write the detailed tab
    df_detailed.to_excel(writer, sheet_name='Detailed', index=False)
    
    # Access the workbook and sheet
    workbook = writer.book
    detailed_sheet = writer.sheets['Detailed']
    
    # Define color coding for severity with fallback for 'warn' and 'fail'
    severity_colors = {
        'info': 'ADD8E6',     # Blue
        'low': 'FFFF00',      # Yellow
        'medium': 'FFA500',   # Orange
        'high': 'FF0000',     # Red
        'warn': 'FFA500',     # Orange like medium
        'fail': 'FF0000'      # Red like high
    }
    
    # Apply color coding based on severity
    for row in range(2, len(df_detailed) + 2):
        severity = detailed_sheet.cell(row=row, column=5).value
        fill_color = severity_colors.get(severity, 'FFFFFF')  # Default to white if not found
        for col in range(1, 6):  # Assuming columns A-E have relevant data
            detailed_sheet.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Auto-filter and auto-width for columns
    detailed_sheet.auto_filter.ref = detailed_sheet.dimensions
    for col in detailed_sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        detailed_sheet.column_dimensions[column].width = adjusted_width
    
    # Create another tab if needed (for example, a summary)
    summary_data = df_detailed.groupby(['Target', 'Audit']).size().unstack(fill_value=0)
    summary_data.to_excel(writer, sheet_name='Summary')

print(f"Report generated successfully: {output_file}")
